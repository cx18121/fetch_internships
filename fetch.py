"""
Summer 2026 Internship Fetcher
==============================
Clones SimplifyJobs/Summer2026-Internships, parses today's (0d) postings,
filters for US-only SWE + DS/AI/ML roles, and saves a formatted Excel file.

Requirements:
    pip install openpyxl beautifulsoup4

Usage:
    python fetch_internships.py

Output:
    Desktop\\Summer2026_Internships_YYYY-MM-DD.xlsx  (Windows)
    ~/Desktop/Summer2026_Internships_YYYY-MM-DD.xlsx (Mac/Linux)
"""

from fileinput import filename
import argparse
import subprocess
import sys
import os
import re
import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# 0. Auto-install dependencies
# ---------------------------------------------------------------------------
def pip_install(*packages):
    subprocess.check_call([sys.executable, "-m", "pip", "install", *packages, "-q"])

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("Installing beautifulsoup4...")
    pip_install("beautifulsoup4")
    from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    pip_install("openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. Clone / update the repo
# ---------------------------------------------------------------------------
REPO_URL = "https://github.com/SimplifyJobs/Summer2026-Internships.git"
REPO_DIR = Path(__file__).parent / "Summer2026-Internships"

def clone_repo():
    if REPO_DIR.exists():
        print("Repo already exists — fetching latest...")
        subprocess.run(["git", "-C", str(REPO_DIR), "fetch", "--depth", "1", "origin"], check=True)
        subprocess.run(["git", "-C", str(REPO_DIR), "reset", "--hard", "origin/HEAD"], check=True)
    else:
        print("Cloning repo (shallow)...")
        subprocess.run(["git", "clone", REPO_URL, str(REPO_DIR), "--depth", "1"], check=True)

# ---------------------------------------------------------------------------
# 2. Parse README.md
# ---------------------------------------------------------------------------
SECTIONS = {
    "## 💻 Software Engineering": "Software Engineering",
    "## 🤖 Data Science":         "Data Science / AI / ML",
}

EXCLUDE_LOCATIONS = [
    "canada","uk","united kingdom","ireland","belfast","toronto","ontario","ottawa",
    "germany","france","india","japan","australia","netherlands","sweden","switzerland",
    "singapore","china","new zealand","austria","denmark","finland","norway","brussels",
    "london","munich","berlin","amsterdam","longueuil","remote in canada",
]

def strip_tags(html):
    return BeautifulSoup(html, "html.parser").get_text(separator=" ").strip()

def first_href(td):
    a = td.find("a")
    if a and a.get("href"):
        href = a["href"]
        # Prefer direct job-board link over simplify redirect
        if "simplify.jobs" not in href:
            return href
        # Fall back to any other <a> in the cell
        for tag in td.find_all("a"):
            h = tag.get("href", "")
            if h and "simplify.jobs" not in h:
                return h
        return href  # last resort: simplify link
    return ""

def parse_readme(days_back=0):
    readme = (REPO_DIR / "README.md").read_text(encoding="utf-8")

    # Split into lines to find section boundaries
    lines = readme.splitlines()
    sections_raw = {}
    current_key = None
    current_lines = []

    for line in lines:
        for prefix, label in SECTIONS.items():
            if line.startswith(prefix):
                if current_key:
                    sections_raw[current_key] = "\n".join(current_lines)
                current_key = label
                current_lines = [line]
                break
        else:
            if current_key:
                # Stop collecting when we hit another ## heading (different section)
                if line.startswith("## ") and not any(line.startswith(p) for p in SECTIONS):
                    sections_raw[current_key] = "\n".join(current_lines)
                    current_key = None
                    current_lines = []
                else:
                    current_lines.append(line)

    if current_key:
        sections_raw[current_key] = "\n".join(current_lines)

    internships = []

    for category, md_chunk in sections_raw.items():
        # Convert markdown to HTML for BeautifulSoup parsing
        soup = BeautifulSoup(md_chunk, "html.parser")
        tables = soup.find_all("table")
        if not tables:
            # Try parsing raw HTML table tags embedded in the markdown
            soup2 = BeautifulSoup(md_chunk.replace("<!--", "<!-- "), "html.parser")
            tables = soup2.find_all("table")

        prev_company = ""
        for table in tables:
            for tr in table.find_all("tr"):
                tds = tr.find_all("td")
                if len(tds) < 5:
                    continue

                # Company
                company_td = tds[0]
                company_text = strip_tags(str(company_td))
                if company_text.strip() in ("↳", ""):
                    company = prev_company
                else:
                    a_tag = company_td.find("a")
                    company = a_tag.get_text(strip=True) if a_tag else company_text.strip()
                    prev_company = company

                # Role
                role = strip_tags(str(tds[1]))

                # Location
                location = strip_tags(str(tds[2]))

                # Apply URL
                apply_url = first_href(tds[3])

                # Age
                age = strip_tags(str(tds[4]))

                age_str = age.strip()
                m = re.match(r"^(\d+)d$", age_str)
                if not m or int(m.group(1)) > days_back:
                    continue

                internships.append({
                    "company":    company,
                    "role":       role,
                    "category":  category,
                    "location":  location,
                    "apply_url": apply_url,
                    "age":       age,
                })

    return internships

# ---------------------------------------------------------------------------
# 3 & 4. Filter: US only + no grad-degree emoji
# ---------------------------------------------------------------------------
def is_us_location(loc):
    loc_lower = loc.lower()
    return not any(term in loc_lower for term in EXCLUDE_LOCATIONS)

def no_grad_required(role):
    return "🎓" not in role

def filter_internships(internships):
    stats = {"non_us": 0, "grad_required": 0}
    kept = []
    for job in internships:
        if not is_us_location(job["location"]):
            stats["non_us"] += 1
            continue
        if not no_grad_required(job["role"]):
            stats["grad_required"] += 1
            continue
        kept.append(job)
    return kept, stats

# ---------------------------------------------------------------------------
# 5. Scoring — tailored to Charles Xue's resume
#
# Profile summary:
#   Cornell CS + Statistics (4.0 GPA, graduating 2028)
#   Languages  : Python, TypeScript, JavaScript, Java, Go, SQL, Bash, R
#   Frameworks : React, Next.js, Node.js, FastAPI, Flask, Javalin, Electron
#   Infra/Tools: PostgreSQL, Redis, Docker, Supabase, Celery, Git, Linux
#   AI/LLM     : Claude Agent SDK, Anthropic API, Gemini API
#   Security   : CTF top-100 US, reverse engineering, C2 implants, Ghidra
#   Project DNA: full-stack AI products, distributed systems, REST APIs,
#                data pipelines, real-time dashboards, desktop apps
# ---------------------------------------------------------------------------

# Tier 1 (+4): Roles that map almost perfectly to Charles's project work
TIER1_ROLE = [
    "software engineer", "software development", "swe intern",
    "full stack", "fullstack", "full-stack",
    "backend", "back end", "back-end",
    "frontend", "front end", "front-end",
    "ai engineer", "ai/ml", "ml engineer",
    "generative ai", "llm", "applied ai",
    "security engineer", "security intern", "cybersecurity",
    "platform engineer", "systems engineer",
]

# Tier 2 (+3): Strong overlap with his stack or interests
TIER2_ROLE = [
    "data engineer", "data science", "machine learning",
    "distributed systems", "infrastructure", "cloud engineer",
    "devops", "site reliability", "sre",
    "api", "rest", "microservices",
    "web developer", "web engineer",
    "embedded", "systems programming",
    "network security", "reverse engineer",
    "quantitative", "quant",          # stats major relevance
]

# Tier 3 (+2): Adjacent / transferable
TIER3_ROLE = [
    "product engineer", "developer", "programmer",
    "data analyst", "analytics engineer",
    "research engineer", "research scientist",
    "mobile", "ios", "android",        # minor overlap, has JS/TS
    "tooling", "build engineer",
    "computer vision", "nlp",
    "blockchain",                       # adjacent to distributed systems
]

# Language / tech keywords (+2 each) — direct resume matches
TECH_STACK = [
    "python", "typescript", "javascript", "java", "golang", "go ",
    "react", "next.js", "nextjs", "node", "node.js",
    "fastapi", "flask", "postgresql", "postgres", "redis",
    "docker", "kubernetes", "aws", "gcp", "azure",
    "celery", "supabase", "graphql",
    "rust",                             # adjacent to systems interests
    "llm", "openai", "anthropic", "langchain",
]

# Company-type signals (+1 each) — Charles's target environment
COMPANY_SIGNALS = [
    "ai", "startup", "fintech", "infrastructure", "developer tools",
    "devtools", "platform", "open source", "cybersecurity", "defense",
    "data", "analytics",
]

# Negative signals (−2 each) — poor fit for a SWE/stats undergrad
NEGATIVE_SIGNALS = [
    "accounting", "finance intern", "marketing", "sales",
    "operations intern", "hr ", "recruiter", "business analyst",
    "hardware", "electrical engineer", "mechanical engineer",
    "civil engineer", "manufacturing",
]

def score(job):
    role_lower    = job["role"].lower()
    company_lower = job["company"].lower()
    full_text     = role_lower + " " + company_lower + " " + job["category"].lower()

    s = 0

    # Tier scoring on role
    for kw in TIER1_ROLE:
        if kw in role_lower:
            s += 4
            break   # only count best tier once per group
    else:
        for kw in TIER2_ROLE:
            if kw in role_lower:
                s += 3
                break
        else:
            for kw in TIER3_ROLE:
                if kw in role_lower:
                    s += 2
                    break

    # Tech stack hits (cumulative)
    s += sum(2 for kw in TECH_STACK if kw in full_text)

    # Company/domain signals (cumulative, capped at +4)
    company_bonus = sum(1 for kw in COMPANY_SIGNALS if kw in full_text)
    s += min(company_bonus, 4)

    # Penalties
    s -= sum(2 for kw in NEGATIVE_SIGNALS if kw in role_lower)

    return max(s, 0)   # floor at 0

def score_label(s):
    if s >= 10: return "⭐⭐⭐ Excellent"
    if s >= 6:  return "⭐⭐ Good"
    if s >= 3:  return "⭐ Decent"
    return "— Low"

# ---------------------------------------------------------------------------
# 6. Build Excel
# ---------------------------------------------------------------------------
NAVY   = "1E3A5F"
WHITE  = "FFFFFF"
BLUE   = "1155CC"
GRAY   = "BBBBBB"
LIGHT  = "EEF2F7"

def thin_border():
    s = Side(style="thin", color=GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def build_excel(internships, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summer 2026 Internships"

    # Header
    headers = ["#", "Company", "Role", "Category", "Location", "Fit", "Score", "Apply"]
    col_widths = [4, 22, 40, 20, 22, 16, 7, 10]

    header_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    header_fill  = PatternFill("solid", start_color=NAVY)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    base_font = Font(name="Arial", size=9)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")

    # Tier fill colors
    TIER_FILLS = {
        "⭐⭐⭐ Excellent": PatternFill("solid", start_color="D6EAD0"),  # soft green
        "⭐⭐ Good":        PatternFill("solid", start_color="FFF9C4"),  # soft yellow
        "⭐ Decent":       PatternFill("solid", start_color="FCE4D6"),  # soft orange
        "— Low":           PatternFill("solid", start_color="F2F2F2"),  # light gray
    }

    for row_idx, job in enumerate(internships, start=2):
        label = score_label(job["score"])
        tier_fill = TIER_FILLS[label]
        use_alt   = (row_idx % 2 == 0)

        def wr(col, value, align=None, font=None, custom_fill=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font   = font or base_font
            c.border = thin_border()
            c.alignment = align or Alignment(vertical="top")
            c.fill = custom_fill or tier_fill

        wr(1, row_idx - 1, center_align)
        wr(2, job["company"])
        wr(3, job["role"], wrap_align)
        wr(4, job["category"])
        wr(5, job["location"], wrap_align)
        wr(6, label, center_align)
        wr(7, job["score"], center_align)

        # Apply hyperlink cell
        apply_cell = ws.cell(row=row_idx, column=8, value="Apply →")
        if job["apply_url"]:
            apply_cell.hyperlink = job["apply_url"]
        apply_cell.font = Font(name="Arial", size=9, color=BLUE, underline="single")
        apply_cell.alignment = center_align
        apply_cell.border = thin_border()
        apply_cell.fill = tier_fill

        ws.row_dimensions[row_idx].height = 36

    # Legend sheet
    lg = wb.create_sheet("Legend")
    lg["A1"] = "Summer 2026 Internship Tracker — Scoring Legend"
    lg["A1"].font = Font(name="Arial", bold=True, size=12, color=NAVY)

    lg["A3"] = "Row Color"
    lg["B3"] = "Fit Label"
    lg["C3"] = "Score Range"
    lg["D3"] = "Meaning"
    for col in ["A3","B3","C3","D3"]:
        lg[col].font = Font(bold=True)

    tiers = [
        ("Green",  "⭐⭐⭐ Excellent", "10+",  "Strong match — role + tech stack align closely with your projects"),
        ("Yellow", "⭐⭐ Good",        "6–9",  "Good fit — relevant role, partial tech overlap"),
        ("Orange", "⭐ Decent",       "3–5",  "Transferable skills, worth a look"),
        ("Gray",   "— Low",           "0–2",  "Weak fit — may not be relevant to your profile"),
    ]
    for i, (color, label, rng, meaning) in enumerate(tiers, start=4):
        lg[f"A{i}"] = color
        lg[f"B{i}"] = label
        lg[f"C{i}"] = rng
        lg[f"D{i}"] = meaning

    lg["A9"]  = "Scoring Breakdown"
    lg["A9"].font = Font(bold=True)
    breakdown = [
        ("Tier 1 role match (+4)", "software engineer, full-stack, AI engineer, security, platform engineer, etc."),
        ("Tier 2 role match (+3)", "data engineer, ML, infra, cloud, DevOps, distributed systems, quant, etc."),
        ("Tier 3 role match (+2)", "developer, analyst, research engineer, mobile, tooling, NLP, etc."),
        ("Tech stack match (+2 each)", "Python, TypeScript, React, Next.js, FastAPI, Go, PostgreSQL, Redis, Docker, LLM, etc."),
        ("Company domain signal (+1, max +4)", "AI, fintech, startup, platform, security, devtools, data, etc."),
        ("Negative signal (−2 each)", "accounting, marketing, sales, HR, hardware, mechanical/electrical engineering, etc."),
    ]
    for i, (a, b) in enumerate(breakdown, start=10):
        lg[f"A{i}"] = a
        lg[f"B{i}"] = b

    lg.column_dimensions["A"].width = 32
    lg.column_dimensions["B"].width = 55
    lg.column_dimensions["C"].width = 14
    lg.column_dimensions["D"].width = 55

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")

# ---------------------------------------------------------------------------
# 7. Email
# ---------------------------------------------------------------------------
def load_env():
    """Load key=value pairs from .env file next to this script into os.environ."""
    env_file = Path(__file__).parent / ".env"
    if not env_file.exists():
        return
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        os.environ.setdefault(key.strip(), value.strip())

def send_email(output_path, to_addr, stats, by_tier, by_cat, label):
    import smtplib
    from email.message import EmailMessage

    load_env()
    gmail_user = os.environ.get("GMAIL_USER")
    gmail_pass = os.environ.get("GMAIL_APP_PASSWORD")
    if not gmail_user or not gmail_pass:
        print("⚠️  Skipping email: set GMAIL_USER and GMAIL_APP_PASSWORD env vars.")
        return

    today = datetime.date.today().strftime("%Y-%m-%d")
    excellent = by_tier.get("⭐⭐⭐ Excellent", 0)
    good      = by_tier.get("⭐⭐ Good", 0)

    body = f"""\
Summer 2026 Internship Digest — {today}

{label} postings found : {sum(by_cat.values()) + stats['non_us'] + stats['grad_required']}
Excluded (non-US)      : {stats['non_us']}
Excluded (grad 🎓)     : {stats['grad_required']}
──────────────────────────────
Kept                   : {sum(by_cat.values())}

By fit tier:
  ⭐⭐⭐ Excellent  {excellent}
  ⭐⭐ Good        {good}
  ⭐ Decent        {by_tier.get('⭐ Decent', 0)}
  — Low            {by_tier.get('— Low', 0)}

By category:
""" + "\n".join(f"  • {cat:<28} {count}" for cat, count in sorted(by_cat.items())) + "\n\nSpreadsheet attached."

    msg = EmailMessage()
    msg["Subject"] = f"[Internships] {excellent} Excellent, {good} Good — {today}"
    msg["From"]    = gmail_user
    msg["To"]      = to_addr
    msg.set_content(body)

    with open(output_path, "rb") as f:
        msg.add_attachment(f.read(),
                           maintype="application",
                           subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           filename=output_path.name)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(gmail_user, gmail_pass)
        smtp.send_message(msg)
    print(f"📧 Email sent to {to_addr}")

# ---------------------------------------------------------------------------
# 8. Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Summer 2026 Internship Fetcher")
    parser.add_argument(
        "--days", "-d",
        type=int,
        default=0,
        metavar="N",
        help="Include postings up to N days old (default: 0 = today only)",
    )
    parser.add_argument(
        "--email", "-e",
        metavar="ADDRESS",
        help="Send the spreadsheet to this email address (requires GMAIL_USER and GMAIL_APP_PASSWORD env vars)",
    )
    args = parser.parse_args()
    days_back = args.days

    today = datetime.date.today().strftime("%Y-%m-%d")
    filename = f"Summer2026_Internships_{today}.xlsx"

    out_dir = Path(__file__).parent / "Internships"
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / filename

    print("=" * 60)
    print("  Summer 2026 Internship Fetcher")
    print("=" * 60)
    if days_back > 0:
        print(f"  Looking back {days_back} day(s)")

    # Step 1: Clone
    clone_repo()

    # Step 2: Parse
    print("\nParsing README.md...")
    all_jobs = parse_readme(days_back)
    label = f"0–{days_back}d" if days_back > 0 else "0d"
    print(f"  Found {len(all_jobs)} total {label} postings")

    if not all_jobs:
        print(f"\n⚠️  No postings found in the last {days_back} day(s). Try a larger --days value or check back later.")
        return

    # Step 3 & 4: Filter
    filtered, stats = filter_internships(all_jobs)
    excluded_total = len(all_jobs) - len(filtered)

    # Step 5: Score & sort
    for job in filtered:
        job["score"] = score(job)
    filtered.sort(key=lambda x: x["score"], reverse=True)

    # Step 6: Build Excel
    print(f"\nBuilding Excel with {len(filtered)} internships...")
    build_excel(filtered, output_path)

    # Step 7: Report
    by_cat = {}
    by_tier = {}
    for job in filtered:
        by_cat[job["category"]] = by_cat.get(job["category"], 0) + 1
        lbl = score_label(job["score"])
        by_tier[lbl] = by_tier.get(lbl, 0) + 1

    print("\n" + "=" * 60)
    print("  RESULTS SUMMARY")
    print("=" * 60)
    print(f"  Total {label} postings found  : {len(all_jobs)}")
    print(f"  Excluded (non-US location) : {stats['non_us']}")
    print(f"  Excluded (grad degree 🎓)  : {stats['grad_required']}")
    print(f"  ─────────────────────────────")
    print(f"  Final internships kept     : {len(filtered)}")
    print(f"\n  By category:")
    for cat, count in sorted(by_cat.items()):
        print(f"    • {cat:<28} {count}")
    print(f"\n  By fit tier:")
    for lbl in ["⭐⭐⭐ Excellent", "⭐⭐ Good", "⭐ Decent", "— Low"]:
        count = by_tier.get(lbl, 0)
        print(f"    {lbl:<20} {count}")
    print(f"\n  📄 File saved to:")
    print(f"     {output_path}")
    print("=" * 60)

    if args.email:
        send_email(output_path, args.email, stats, by_tier, by_cat, label)

if __name__ == "__main__":
    main()